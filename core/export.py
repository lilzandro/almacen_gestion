import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl.utils
from database.repository import get_all_movements, get_all_products


def _header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill("solid", fgColor="003366"),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            *[Side(style="thin")] * 0,
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


def _apply_headers(ws, headers, col_widths):
    h = _header_style()
    for col, (text, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = h["font"]
        cell.fill = h["fill"]
        cell.alignment = h["alignment"]
        cell.border = h["border"]
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def _row_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def export_movements(filepath: str):
    movements = get_all_movements(limit=10000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    alt = PatternFill("solid", fgColor="EEF2FF")
    border = _row_border()

    headers = [
        "ID",
        "Tipo",
        "Fecha/Hora",
        "Cantidad",
        "Producto",
        "Empleado",
        "Registrado por",
        "Notas",
    ]
    widths = [6, 12, 18, 8, 32, 22, 16, 28]
    _apply_headers(ws, headers, widths)

    for r, row in enumerate(movements, 2):
        data = [
            row["id"],
            row["type"],
            row["timestamp"],
            row["quantity"],
            row["product"],
            row["employee"],
            row["registered_by"],
            row["notes"],
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if r % 2 == 0:
                cell.fill = alt
    wb.save(filepath)


def export_inventory(filepath: str):
    products = get_all_products()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    border = _row_border()
    status_colors = {
        "disponible": "C6EFCE",
        "no disponible": "FFEB9C",
        "inactivo": "FFC7CE",
    }

    headers = [
        "ID",
        "Nombre",
        "Código de Barras",
        "Cantidad",
        "Estado",
        "Proveedor",
        "Ubicación",
    ]
    widths = [6, 26, 20, 12, 16, 26, 16]
    _apply_headers(ws, headers, widths)

    for r, row in enumerate(products, 2):
        data = [
            row["id"],
            row["name"],
            row["barcode"],
            row["quantity"],
            row["status"],
            row["supplier_name"],
            row["location"],
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if c == 5 and row["status"] in status_colors:
                cell.fill = PatternFill("solid", fgColor=status_colors[row["status"]])
    wb.save(filepath)
